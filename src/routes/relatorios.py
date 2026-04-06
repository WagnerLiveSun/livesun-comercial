import io
from collections import defaultdict
from types import SimpleNamespace
from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from datetime import datetime
from flask_login import login_required, current_user
from src.models import db, Lancamento, Entidade, ContaBanco, FluxoContaModel, ImportacaoNFSe, ConciliacaoItem
from sqlalchemy.orm import joinedload
from sqlalchemy import func, or_
from decimal import Decimal
import logging

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None
    logging.getLogger(__name__).warning("openpyxl not available; Excel exports disabled", exc_info=True)

relatorios_bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')


def _descricao_lancamento_fluxo(lancamento: Lancamento) -> str:
    """Retorna descrição amigável para relatórios de fluxo.

    Prioridade:
    1) Observações (importações NFSe já gravam descrição do serviço aqui)
    2) Nome da entidade (para lançamentos manuais sem observação)
    3) Placeholder
    """
    observacoes = (getattr(lancamento, 'observacoes', None) or '').strip()
    if observacoes:
        return observacoes

    if getattr(lancamento, 'entidade', None) and lancamento.entidade.nome:
        return lancamento.entidade.nome

    return '-'


def _build_consolidado_por_fluxo(lancamentos):
    """Consolida lançamentos por conta de fluxo para visão gerencial."""
    agrupado = {}

    for lancamento in lancamentos:
        fluxo = lancamento.fluxo_conta
        fluxo_id = getattr(fluxo, 'id', 0) or 0
        codigo = getattr(fluxo, 'codigo', None) or '-'
        descricao = getattr(fluxo, 'descricao', None) or 'Sem conta de fluxo'
        tipo = fluxo.get_tipo_descricao() if fluxo else '-'

        if fluxo_id not in agrupado:
            agrupado[fluxo_id] = {
                'conta_fluxo': f"{codigo} - {descricao}" if codigo != '-' else descricao,
                'tipo': tipo,
                'qtd_lancamentos': 0,
                'total_previsto_pagar': Decimal('0.00'),
                'total_previsto_receber': Decimal('0.00'),
                'total_realizado_pagar': Decimal('0.00'),
                'total_realizado_receber': Decimal('0.00'),
            }

        row = agrupado[fluxo_id]
        row['qtd_lancamentos'] += 1

        valor_real = Decimal(str(lancamento.valor_real or 0))
        valor_pago = Decimal(str(lancamento.valor_pago or 0))

        if fluxo and fluxo.is_pagamento():
            row['total_previsto_pagar'] += valor_real
            row['total_realizado_pagar'] += valor_pago
        else:
            row['total_previsto_receber'] += valor_real
            row['total_realizado_receber'] += valor_pago

    rows = []
    for row in agrupado.values():
        row['saldo_previsto'] = row['total_previsto_receber'] - row['total_previsto_pagar']
        row['saldo_realizado'] = row['total_realizado_receber'] - row['total_realizado_pagar']
        row['total_previsto'] = row['total_previsto_receber'] + row['total_previsto_pagar']
        row['total_realizado'] = row['total_realizado_receber'] + row['total_realizado_pagar']
        rows.append(SimpleNamespace(**row))

    rows.sort(key=lambda x: x.conta_fluxo)
    return rows

# --- LISTAGEM DE NOTAS EMITIDAS (NFSe) ---
@relatorios_bp.route('/notas_nfse', methods=['GET'])
@login_required
def listagem_notas_nfse():
    nome_cliente = request.args.get('nome_cliente', '').strip()
    data_emissao_de = request.args.get('data_emissao_de', '')
    data_emissao_ate = request.args.get('data_emissao_ate', '')
    data_venc_de = request.args.get('data_venc_de', '')
    data_venc_ate = request.args.get('data_venc_ate', '')
    status = request.args.get('status', '')

    query = ImportacaoNFSe.query.filter(ImportacaoNFSe.empresa_id == current_user.empresa_id)
    query = query.options(joinedload(ImportacaoNFSe.entidade), joinedload(ImportacaoNFSe.lancamento))

    if nome_cliente:
        query = query.join(Entidade, ImportacaoNFSe.entidade).filter(Entidade.nome.ilike(f"%{nome_cliente}%"))
    if data_emissao_de:
        try:
            data_emissao_de_dt = datetime.strptime(data_emissao_de, '%Y-%m-%d').date()
            query = query.filter(ImportacaoNFSe.data_emissao >= data_emissao_de_dt)
        except Exception:
            flash('Data de emissão (de) inválida.', 'warning')
    if data_emissao_ate:
        try:
            data_emissao_ate_dt = datetime.strptime(data_emissao_ate, '%Y-%m-%d').date()
            query = query.filter(ImportacaoNFSe.data_emissao <= data_emissao_ate_dt)
        except Exception:
            flash('Data de emissão (até) inválida.', 'warning')
    if data_venc_de:
        try:
            data_venc_de_dt = datetime.strptime(data_venc_de, '%Y-%m-%d').date()
            query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.data_vencimento >= data_venc_de_dt)
        except Exception:
            flash('Data de vencimento (de) inválida.', 'warning')
    if data_venc_ate:
        try:
            data_venc_ate_dt = datetime.strptime(data_venc_ate, '%Y-%m-%d').date()
            query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.data_vencimento <= data_venc_ate_dt)
        except Exception:
            flash('Data de vencimento (até) inválida.', 'warning')
    if status:
        query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.status == status)

    notas = query.order_by(ImportacaoNFSe.data_emissao.desc(), ImportacaoNFSe.id.desc()).all()

    total_valor_bruto = sum([float(n.valor_bruto or 0) for n in notas])
    total_valor_impostos = sum([
        float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0)
        for n in notas
    ])

    return render_template(
        'relatorios/listagem_notas_nfse.html',
        notas=notas,
        total_valor_bruto=total_valor_bruto,
        total_valor_impostos=total_valor_impostos,
        nome_cliente=nome_cliente,
        data_emissao_de=data_emissao_de,
        data_emissao_ate=data_emissao_ate,
        data_venc_de=data_venc_de,
        data_venc_ate=data_venc_ate,
        status=status,
        empresa_nome=(current_user.empresa.nome if current_user.empresa else '-'),
        empresa_cnpj=(current_user.empresa.cnpj if current_user.empresa else '-'),
        gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'),
        get_valor_imposto=lambda n: float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0)
    )


# Exportação Excel/PDF
@relatorios_bp.route('/notas_nfse/export', methods=['GET'])
@login_required
def export_listagem_notas_nfse():
    formato = request.args.get('formato', 'xlsx').lower()
    nome_cliente = request.args.get('nome_cliente', '').strip()
    data_emissao_de = request.args.get('data_emissao_de', '')
    data_emissao_ate = request.args.get('data_emissao_ate', '')
    data_venc_de = request.args.get('data_venc_de', '')
    data_venc_ate = request.args.get('data_venc_ate', '')
    status = request.args.get('status', '')

    query = ImportacaoNFSe.query.filter(ImportacaoNFSe.empresa_id == current_user.empresa_id)
    query = query.options(joinedload(ImportacaoNFSe.entidade), joinedload(ImportacaoNFSe.lancamento))
    if nome_cliente:
        query = query.join(Entidade, ImportacaoNFSe.entidade).filter(Entidade.nome.ilike(f"%{nome_cliente}%"))
    if data_emissao_de:
        try:
            data_emissao_de_dt = datetime.strptime(data_emissao_de, '%Y-%m-%d').date()
            query = query.filter(ImportacaoNFSe.data_emissao >= data_emissao_de_dt)
        except Exception:
            pass
    if data_emissao_ate:
        try:
            data_emissao_ate_dt = datetime.strptime(data_emissao_ate, '%Y-%m-%d').date()
            query = query.filter(ImportacaoNFSe.data_emissao <= data_emissao_ate_dt)
        except Exception:
            pass
    if data_venc_de:
        try:
            data_venc_de_dt = datetime.strptime(data_venc_de, '%Y-%m-%d').date()
            query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.data_vencimento >= data_venc_de_dt)
        except Exception:
            pass
    if data_venc_ate:
        try:
            data_venc_ate_dt = datetime.strptime(data_venc_ate, '%Y-%m-%d').date()
            query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.data_vencimento <= data_venc_ate_dt)
        except Exception:
            pass
    if status:
        query = query.join(Lancamento, ImportacaoNFSe.lancamento).filter(Lancamento.status == status)

    notas = query.order_by(ImportacaoNFSe.data_emissao.desc(), ImportacaoNFSe.id.desc()).all()

    if formato == 'xlsx':
        if Workbook is None:
            flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
            return redirect(url_for('relatorios.listagem_notas_nfse', **request.args))
        wb = Workbook()
        ws = wb.active
        ws.title = 'Notas NFSe'
        ws.append(['Empresa', current_user.empresa.nome if current_user.empresa else '-'])
        ws.append(['CNPJ', current_user.empresa.cnpj if current_user.empresa else '-'])
        ws.append(['Gerado em', datetime.now().strftime('%d/%m/%Y %H:%M')])
        ws.append([])
        ws.append([
            'ID', 'Empresa ID', 'Número Nota', 'Data Emissão', 'CNPJ Tomador',
            'Cliente', 'Valor Bruto', 'Alíquota ISS', 'Valor Impostos', 'Status', 'Data Vencimento', 'Data Pagamento', 'Descrição Serviço'
        ])
        for n in notas:
            valor_imposto = float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0)
            ws.append([
                n.id,
                n.empresa_id,
                n.numero_nota or '',
                n.data_emissao.strftime('%d/%m/%Y') if n.data_emissao else '',
                n.cnpj_tomador or '',
                (n.entidade.nome if n.entidade else ''),
                float(n.valor_bruto or 0),
                float(n.aliquota_iss or 0),
                valor_imposto,
                (n.lancamento.status if n.lancamento else ''),
                n.lancamento.data_vencimento.strftime('%d/%m/%Y') if n.lancamento and n.lancamento.data_vencimento else '',
                n.lancamento.data_pagamento.strftime('%d/%m/%Y') if n.lancamento and n.lancamento.data_pagamento else '',
                n.descricao_servico or ''
            ])
        ws.append([])
        ws.append([
            'TOTAIS', '', '', '', '', '',
            sum([float(n.valor_bruto or 0) for n in notas]),
            '',
            sum([
                float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0)
                for n in notas
            ]), '', '', '', ''
        ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='listagem_notas_nfse.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if formato == 'pdf':
        try:
            from fpdf import FPDF
        except ImportError:
            flash('Exportação para PDF indisponível: biblioteca "fpdf" não está instalada no ambiente.', 'warning')
            return redirect(url_for('relatorios.listagem_notas_nfse', **request.args))

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(0, 7, 'Listagem de Notas NFSe', ln=1)
        pdf.set_font('Arial', '', 8)
        pdf.cell(0, 5, f"Empresa: {current_user.empresa.nome if current_user.empresa else '-'} | CNPJ: {current_user.empresa.cnpj if current_user.empresa else '-'}", ln=1)
        pdf.cell(0, 5, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=1)
        pdf.ln(1)

        headers = [
            ('ID', 8), ('Nº Nota', 12), ('Data Emissão', 14), ('CNPJ Tomador', 20),
            ('Cliente', 22), ('Vlr Bruto', 14), ('Aliq. ISS', 10), ('Impostos', 14),
            ('Status', 12), ('Dt Venc.', 14), ('Dt Pag.', 14), ('Chave Nota', 60), ('Descrição', 60)
        ]

        pdf.set_font('Arial', 'B', 5)
        for titulo, largura in headers:
            pdf.cell(largura, 5, titulo, 1)
        pdf.ln()

        pdf.set_font('Arial', '', 5)
        for n in notas:
            valor_imposto = float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0)
            pdf.cell(8, 4, str(n.id), 1)
            pdf.cell(12, 4, str(n.numero_nota or ''), 1)
            pdf.cell(14, 4, n.data_emissao.strftime('%d/%m/%Y') if n.data_emissao else '-', 1)
            pdf.cell(20, 4, str(n.cnpj_tomador or ''), 1)
            pdf.cell(22, 4, (n.entidade.nome if n.entidade else '-')[:14], 1)
            pdf.cell(14, 4, f"{float(n.valor_bruto or 0):,.2f}", 1, align='R')
            pdf.cell(10, 4, f"{float(n.aliquota_iss) if n.aliquota_iss is not None else '-'}", 1, align='R')
            pdf.cell(14, 4, f"{valor_imposto:,.2f}", 1, align='R')
            pdf.cell(12, 4, str(n.lancamento.status if n.lancamento else '-'), 1)
            pdf.cell(14, 4, n.lancamento.data_vencimento.strftime('%d/%m/%Y') if n.lancamento and n.lancamento.data_vencimento else '-', 1)
            pdf.cell(14, 4, n.lancamento.data_pagamento.strftime('%d/%m/%Y') if n.lancamento and n.lancamento.data_pagamento else '-', 1)
            pdf.cell(60, 4, str(n.chave_nota or ''), 1)
            pdf.cell(60, 4, (n.descricao_servico or '-')[:40], 1)
            pdf.ln()

        pdf.set_font('Arial', 'B', 5)
        pdf.cell(8 + 12 + 14 + 20 + 22, 5, 'TOTAIS', 1)
        pdf.cell(14, 5, f"{sum([float(n.valor_bruto or 0) for n in notas]):,.2f}", 1, align='R')
        pdf.cell(10, 5, '', 1)
        pdf.cell(14, 5, f"{sum([float(n.lancamento.valor_imposto) if n.lancamento and n.lancamento.valor_imposto is not None else float(n.valor_impostos or 0) for n in notas]):,.2f}", 1, align='R')
        pdf.cell(12 + 14 + 14 + 60 + 60, 5, '', 1)

        output = io.BytesIO()
        output.write(pdf.output(dest='S').encode('latin1'))
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='listagem_notas_nfse.pdf', mimetype='application/pdf')

    flash('Formato de exportação inválido.', 'warning')
    return redirect(url_for('relatorios.listagem_notas_nfse', **request.args))


def _parse_date_filter(value, field_label):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        flash(f'Data inválida para "{field_label}". Use o formato YYYY-MM-DD.', 'warning')
        return None


def _build_listagem_lancamentos_context(args):
    def get_saldo_inicial_por_conta():
        if conta_fluxo_id:
            contas = ContaBanco.query.filter(
                ContaBanco.empresa_id == current_user.empresa_id,
                ContaBanco.id == conta_fluxo_id
            ).all()
        else:
            contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
        return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

    def get_saldo_inicial_total():
        return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

    data_lanc_de = args.get('data_lanc_de', '')
    data_lanc_ate = args.get('data_lanc_ate', '')
    data_venc_de = args.get('data_venc_de', '')
    data_venc_ate = args.get('data_venc_ate', '')
    tipo_movimento = args.get('tipo_movimento', '')
    entidade_tipo = args.get('entidade_tipo', '')
    entidade_id = args.get('entidade_id', type=int)
    conta_fluxo_id = args.get('conta_fluxo_id', type=int)
    status = args.get('status', '')
    conciliacao_status = args.get('conciliacao_status', '')

    data_lanc_de_dt = _parse_date_filter(data_lanc_de, 'Data de lançamento (de)')
    data_lanc_ate_dt = _parse_date_filter(data_lanc_ate, 'Data de lançamento (até)')
    data_venc_de_dt = _parse_date_filter(data_venc_de, 'Data de vencimento (de)')
    data_venc_ate_dt = _parse_date_filter(data_venc_ate, 'Data de vencimento (até)')

    from sqlalchemy.orm import joinedload as _joinedload
    query = Lancamento.query.filter(Lancamento.empresa_id == current_user.empresa_id).options(_joinedload(Lancamento.entidade))

    if data_lanc_de_dt:
        query = query.filter(Lancamento.data_evento >= data_lanc_de_dt)
    if data_lanc_ate_dt:
        query = query.filter(Lancamento.data_evento <= data_lanc_ate_dt)
    if data_venc_de_dt:
        query = query.filter(Lancamento.data_vencimento >= data_venc_de_dt)
    if data_venc_ate_dt:
        query = query.filter(Lancamento.data_vencimento <= data_venc_ate_dt)
    if tipo_movimento in ('P', 'R'):
        query = query.filter(Lancamento.fluxo_conta.has(FluxoContaModel.tipo == tipo_movimento))
    if entidade_tipo in ('C', 'F'):
        query = query.filter(Lancamento.entidade.has(Entidade.tipo == entidade_tipo))
    if entidade_id:
        query = query.filter(Lancamento.entidade_id == entidade_id)
    if conta_fluxo_id:
        query = query.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
    if status:
        query = query.filter(Lancamento.status == status)

    conciliados_ids = set(
        item.lancamento_id
        for item in ConciliacaoItem.query.filter(
            ConciliacaoItem.empresa_id == current_user.empresa_id,
            ConciliacaoItem.status == 'conciliado',
            ConciliacaoItem.lancamento_id.isnot(None)
        ).all()
    )

    lancamentos = query.order_by(
        Lancamento.data_evento.desc(),
        Lancamento.data_vencimento.desc(),
        Lancamento.id.desc()
    ).all()

    for lancamento in lancamentos:
        lancamento.conciliacao_status = 'conciliado' if lancamento.id in conciliados_ids else 'nao_conciliado'

    if conciliacao_status == 'conciliado':
        lancamentos = [l for l in lancamentos if l.conciliacao_status == 'conciliado']
    elif conciliacao_status == 'nao_conciliado':
        lancamentos = [l for l in lancamentos if l.conciliacao_status == 'nao_conciliado']

    total_pago = Decimal('0.00')
    total_recebido = Decimal('0.00')
    total_valor_imposto = Decimal('0.00')
    total_valor_outros_custos = Decimal('0.00')
    for lancamento in lancamentos:
        tipo = lancamento.fluxo_conta.tipo if lancamento.fluxo_conta else None
        valor_pago = Decimal(str(lancamento.valor_pago or 0))
        valor_recebido = Decimal(str(lancamento.valor_pago or 0))
        if tipo == 'P':
            total_pago += valor_pago
        elif tipo == 'R':
            total_recebido += valor_recebido
        total_valor_imposto += Decimal(str(lancamento.valor_imposto or 0))
        total_valor_outros_custos += Decimal(str(lancamento.valor_outros_custos or 0))

    mostrar_pago = entidade_tipo != 'C'
    mostrar_recebido = entidade_tipo != 'F'

    entidades = Entidade.query.filter_by(
        empresa_id=current_user.empresa_id,
        ativo=True
    ).order_by(Entidade.nome.asc()).all()
    contas_fluxo = FluxoContaModel.query.filter_by(
        empresa_id=current_user.empresa_id,
        ativo=True
    ).order_by(FluxoContaModel.codigo.asc()).all()

    return {
        'lancamentos': lancamentos,
        'entidades': entidades,
        'contas_fluxo': contas_fluxo,
        'total_pago': total_pago,
        'total_recebido': total_recebido,
        'total_valor_imposto': total_valor_imposto,
        'total_valor_outros_custos': total_valor_outros_custos,
        'mostrar_pago': mostrar_pago,
        'mostrar_recebido': mostrar_recebido,
        'empresa_nome': (current_user.empresa.nome if current_user.empresa else '-'),
        'empresa_cnpj': (current_user.empresa.cnpj if current_user.empresa else '-'),
        'data_lanc_de': data_lanc_de,
        'data_lanc_ate': data_lanc_ate,
        'data_venc_de': data_venc_de,
        'data_venc_ate': data_venc_ate,
        'tipo_movimento': tipo_movimento,
        'entidade_tipo': entidade_tipo,
        'entidade_id': entidade_id,
        'conta_fluxo_id': conta_fluxo_id,
        'status': status,
        'conciliacao_status': conciliacao_status,
        'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M')
    }

# --- LISTAGEM DE LANÇAMENTOS ---
@relatorios_bp.route('/lancamentos', methods=['GET'])
@login_required
def listagem_lancamentos():
    context = _build_listagem_lancamentos_context(request.args)
    return render_template('relatorios/listagem_lancamentos.html', **context)


@relatorios_bp.route('/lancamentos/export', methods=['GET'])
@login_required
def export_listagem_lancamentos():
    formato = request.args.get('formato', 'xlsx').lower()
    context = _build_listagem_lancamentos_context(request.args)
    lancamentos = context['lancamentos']

    if formato == 'xlsx':
        if Workbook is None:
            flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
            return redirect(url_for('relatorios.listagem_lancamentos', **request.args))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Listagem Lancamentos'
        ws.append(['Empresa', context['empresa_nome']])
        ws.append(['CNPJ', context['empresa_cnpj']])
        ws.append(['Gerado em', context['gerado_em']])
        ws.append([])
        ws.append([
            'ID', 'Empresa', 'Processo', 'Status', 'Conciliação', 'Data Lancamento', 'Data Vencimento',
            'Data Pagamento', 'Entidade', 'Tipo Entidade', 'Conta Fluxo', 'Conta Banco',
            'Documento', 'Valor Real', 'Valor Pago/Recebido', 'Impostos', 'Outros Custos', 'Observacoes'
        ])

        for l in lancamentos:
            processo = l.fluxo_conta.get_tipo_descricao() if l.fluxo_conta else '-'
            tipo_entidade = l.entidade.get_tipo_descricao() if l.entidade else '-'
            conta_fluxo = f"{l.fluxo_conta.codigo} - {l.fluxo_conta.descricao}" if l.fluxo_conta else '-'
            ws.append([
                l.id,
                l.empresa.nome if l.empresa else '-',
                processo,
                l.status,
                'Conciliado' if getattr(l, 'conciliacao_status', '') == 'conciliado' else 'Não conciliado',
                l.data_evento.strftime('%d/%m/%Y') if l.data_evento else '-',
                l.data_vencimento.strftime('%d/%m/%Y') if l.data_vencimento else '-',
                l.data_pagamento.strftime('%d/%m/%Y') if l.data_pagamento else '-',
                l.entidade.nome if l.entidade else '-',
                tipo_entidade,
                conta_fluxo,
                l.conta_banco.nome if l.conta_banco else '-',
                l.numero_documento or '-',
                '',
                float(l.valor_pago or 0) if l.fluxo_conta and l.fluxo_conta.is_pagamento() else '',
                float(l.valor_pago or 0) if l.fluxo_conta and l.fluxo_conta.is_recebimento() else '',
                float(l.valor_imposto or 0),
                float(l.valor_outros_custos or 0),
                l.observacoes or '-'
            ])

        ws.append([])
        ws.append(['TOTAL', '', '', '', '', '', '', '', '', '', '', '', '', '', float(context['total_pago']), float(context['total_recebido']), float(context['total_valor_imposto']), float(context['total_valor_outros_custos']), ''])

        for col in ['O', 'P', 'Q', 'R']:
            for cell in ws[col]:
                cell.number_format = '#,##0.00'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='listagem_lancamentos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    if formato == 'pdf':
        try:
            from fpdf import FPDF
        except ImportError:
            flash('Exportação para PDF indisponível: biblioteca "fpdf" não está instalada no ambiente.', 'warning')
            return redirect(url_for('relatorios.listagem_lancamentos', **request.args))

        total_valor_real = sum(float(l.valor_real or 0) for l in lancamentos)
        total_valor_pago = sum(float(l.valor_pago or 0) for l in lancamentos)

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('Arial', 'B', 11)
        pdf.cell(0, 8, 'Listagem de Lancamentos', ln=1)
        pdf.set_font('Arial', '', 9)
        pdf.cell(0, 6, f"Empresa: {context['empresa_nome']} | CNPJ: {context['empresa_cnpj']}", ln=1)
        pdf.cell(0, 6, f"Gerado em: {context['gerado_em']}", ln=1)
        pdf.ln(2)

        headers = [
            ('ID', 10), ('Proc.', 20), ('Status', 18), ('Conc.', 16), ('Dt Lanc.', 22), ('Dt Venc.', 22),
            ('Dt Pag.', 22), ('Entidade', 45), ('Conta Fluxo', 45), ('Documento', 24),
            ('Vlr Real', 22), ('Vlr Pago', 22), ('Impostos', 22), ('Outros', 22)
        ]

        pdf.set_font('Arial', 'B', 8)
        for titulo, largura in headers:
            pdf.cell(largura, 7, titulo, 1)
        pdf.ln()

        pdf.set_font('Arial', '', 7)
        for l in lancamentos:
            processo = l.fluxo_conta.get_tipo_descricao() if l.fluxo_conta else '-'
            conta_fluxo = f"{l.fluxo_conta.codigo} - {l.fluxo_conta.descricao}" if l.fluxo_conta else '-'
            pdf.cell(10, 6, str(l.id), 1)
            pdf.cell(20, 6, processo, 1)
            pdf.cell(18, 6, str(l.status or '-'), 1)
            pdf.cell(16, 6, 'Sim' if getattr(l, 'conciliacao_status', '') == 'conciliado' else 'Nao', 1)
            pdf.cell(22, 6, l.data_evento.strftime('%d/%m/%Y') if l.data_evento else '-', 1)
            pdf.cell(22, 6, l.data_vencimento.strftime('%d/%m/%Y') if l.data_vencimento else '-', 1)
            pdf.cell(22, 6, l.data_pagamento.strftime('%d/%m/%Y') if l.data_pagamento else '-', 1)
            pdf.cell(45, 6, (l.entidade.nome if l.entidade else '-')[:26], 1)
            pdf.cell(45, 6, conta_fluxo[:28], 1)
            pdf.cell(24, 6, (l.numero_documento or '-')[:14], 1)
            pdf.cell(22, 6, f"{float(l.valor_real or 0):,.2f}", 1, align='R')
            pdf.cell(22, 6, f"{float(l.valor_pago or 0):,.2f}", 1, align='R')
            pdf.cell(22, 6, f"{float(l.valor_imposto or 0):,.2f}", 1, align='R')
            pdf.cell(22, 6, f"{float(l.valor_outros_custos or 0):,.2f}", 1, align='R')
            pdf.ln()

        pdf.set_font('Arial', 'B', 8)
        pdf.cell(228, 7, 'TOTAIS', 1)
        pdf.cell(22, 7, f"{total_valor_real:,.2f}", 1, align='R')
        pdf.cell(22, 7, f"{total_valor_pago:,.2f}", 1, align='R')
        pdf.cell(22, 7, f"{float(context['total_valor_imposto']):,.2f}", 1, align='R')
        pdf.cell(22, 7, f"{float(context['total_valor_outros_custos']):,.2f}", 1, align='R')

        output = io.BytesIO()
        output.write(pdf.output(dest='S').encode('latin1'))
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='listagem_lancamentos.pdf', mimetype='application/pdf')

    flash('Formato de exportação inválido.', 'warning')
    return redirect(url_for('relatorios.listagem_lancamentos', **request.args))


# Relatório de Fluxo de Caixa CSV
@relatorios_bp.route('/fluxo-caixa-csv')
@login_required
def fluxo_caixa_csv():
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    query = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query = query.filter(or_(Lancamento.data_pagamento >= data_inicio_dt, Lancamento.data_vencimento >= data_inicio_dt))
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query = query.filter(or_(Lancamento.data_pagamento <= data_fim_dt, Lancamento.data_vencimento <= data_fim_dt))
    if hasattr(current_user, 'empresa_id'):
        query = query.filter(Lancamento.empresa_id == current_user.empresa_id)
    query = query.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
    lancamentos = query.order_by(
        FluxoContaModel.descricao.asc(),
        func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc(),
        Lancamento.id.asc()
    ).all()
    dados_csv = _build_consolidado_por_fluxo(lancamentos)
    return render_template('relatorios/fluxo_caixa_csv.html', dados_csv=dados_csv, data_inicio=data_inicio, data_fim=data_fim)


@relatorios_bp.route('/fluxo-caixa-csv/export')
@login_required
def export_fluxo_caixa_csv():
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    if Workbook is None:
        flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
        return redirect(url_for('relatorios.fluxo_caixa'))
    query = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query = query.filter(or_(Lancamento.data_pagamento >= data_inicio_dt, Lancamento.data_vencimento >= data_inicio_dt))
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query = query.filter(or_(Lancamento.data_pagamento <= data_fim_dt, Lancamento.data_vencimento <= data_fim_dt))
    if hasattr(current_user, 'empresa_id'):
        query = query.filter(Lancamento.empresa_id == current_user.empresa_id)
    query = query.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
    lancamentos = query.order_by(
        FluxoContaModel.descricao.asc(),
        func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc(),
        Lancamento.id.asc()
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fluxo de Caixa'
    ws.append([
        'Conta Fluxo', 'Tipo', 'Qtde Lançamentos',
        'Previsto a Pagar (R$)', 'Previsto a Receber (R$)', 'Saldo Previsto (R$)',
        'Realizado Pago (R$)', 'Realizado Recebido (R$)', 'Saldo Realizado (R$)'
    ])

    for row in _build_consolidado_por_fluxo(lancamentos):
        ws.append([
            row.conta_fluxo,
            row.tipo,
            row.qtd_lancamentos,
            float(row.total_previsto_pagar),
            float(row.total_previsto_receber),
            float(row.saldo_previsto),
            float(row.total_realizado_pagar),
            float(row.total_realizado_receber),
            float(row.saldo_realizado),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='fluxo_caixa.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@relatorios_bp.route('/fluxo-caixa-previsto')
@login_required
def fluxo_caixa_previsto():
    return render_template('relatorios/fluxo_caixa_previsto.html')


@relatorios_bp.route('/fluxo-caixa-realizado')
@login_required
def fluxo_caixa_realizado():
    return render_template('relatorios/fluxo_caixa_realizado.html')


@relatorios_bp.route('/fluxo-caixa-legacy')
@login_required
def fluxo_caixa():
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    conta_banco_id = request.args.get('conta_banco_id', '', type=int)
    conta_fluxo_id = request.args.get('conta_fluxo_id', '', type=int)

    def get_saldo_inicial_por_conta():
        if conta_banco_id:
            contas = ContaBanco.query.filter(
                ContaBanco.empresa_id == current_user.empresa_id,
                ContaBanco.id == conta_banco_id
            ).all()
        else:
            contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
        return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

    def get_saldo_inicial_total():
        return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

    def build_fluxo_rows(lancamentos, saldo_inicial_por_conta, use_valor_real):
        saldo_atual_por_conta = saldo_inicial_por_conta.copy()
        rows = []
        for lancamento in lancamentos:
            conta_id = lancamento.conta_banco_id
            saldo_anterior = saldo_atual_por_conta.get(conta_id, Decimal('0.00'))
            valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
            valor = Decimal(str(valor_base or 0))
            if lancamento.fluxo_conta and lancamento.fluxo_conta.is_pagamento():
                saldo_atual = saldo_anterior - valor
            else:
                saldo_atual = saldo_anterior + valor
            saldo_atual_por_conta[conta_id] = saldo_atual
            rows.append(SimpleNamespace(
                lancamento=lancamento,
                saldo_anterior=saldo_anterior,
                saldo_atual=saldo_atual
            ))
        return rows

    def build_daily_rows(lancamentos, saldo_inicial, use_valor_real, date_attr):
        totals = defaultdict(lambda: {'pagar': Decimal('0.00'), 'receber': Decimal('0.00')})
        for lancamento in lancamentos:
            data_ref = getattr(lancamento, date_attr)
            if not data_ref:
                continue
            valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
            valor = Decimal(str(valor_base or 0))
            if lancamento.fluxo_conta and lancamento.fluxo_conta.is_pagamento():
                totals[data_ref]['pagar'] += valor
            else:
                totals[data_ref]['receber'] += valor

        rows = []
        saldo_anterior = saldo_inicial
        for data in sorted(totals.keys()):
            pagamentos = totals[data]['pagar']
            recebimentos = totals[data]['receber']
            saldo_atual = saldo_anterior - pagamentos + recebimentos
            rows.append(SimpleNamespace(
                data=data,
                saldo_anterior=saldo_anterior,
                pagamentos=pagamentos,
                recebimentos=recebimentos,
                saldo_atual=saldo_atual
            ))
            saldo_anterior = saldo_atual
        return rows

    saldo_inicial_por_conta = get_saldo_inicial_por_conta()
    saldo_inicial_total = get_saldo_inicial_total()

    query_realizado = Lancamento.query.filter(
        Lancamento.empresa_id == current_user.empresa_id,
        Lancamento.status == 'pago'
    )
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query_realizado = query_realizado.filter(Lancamento.data_pagamento >= data_inicio_dt)
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query_realizado = query_realizado.filter(Lancamento.data_pagamento <= data_fim_dt)
    if conta_banco_id:
        query_realizado = query_realizado.filter(Lancamento.conta_banco_id == conta_banco_id)
    if conta_fluxo_id:
        query_realizado = query_realizado.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
    query_realizado = query_realizado.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
    lancamentos_realizado = query_realizado.order_by(
        FluxoContaModel.descricao.asc(),
        Lancamento.data_pagamento.asc(),
        Lancamento.id.asc()
    ).all()

    query_previsto = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
    if data_inicio:
        query_previsto = query_previsto.filter(Lancamento.data_vencimento >= data_inicio_dt)
    if data_fim:
        query_previsto = query_previsto.filter(Lancamento.data_vencimento <= data_fim_dt)
    if conta_banco_id:
        query_previsto = query_previsto.filter(Lancamento.conta_banco_id == conta_banco_id)
    if conta_fluxo_id:
        query_previsto = query_previsto.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
    query_previsto = query_previsto.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
    lancamentos_previsto = query_previsto.order_by(
        FluxoContaModel.descricao.asc(),
        Lancamento.data_vencimento.asc(),
        Lancamento.id.asc()
    ).all()

    resumo_diario_realizado = build_daily_rows(
        lancamentos_realizado,
        saldo_inicial_total,
        use_valor_real=False,
        date_attr='data_pagamento'
    )

    resumo_diario_previsto = build_daily_rows(
        lancamentos_previsto,
        saldo_inicial_total,
        use_valor_real=True,
        date_attr='data_vencimento'
    )

    consolidado_realizado = _build_consolidado_por_fluxo(lancamentos_realizado)
    consolidado_previsto = _build_consolidado_por_fluxo(lancamentos_previsto)

    contas_banco = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
    contas_fluxo = FluxoContaModel.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()

    return render_template(
        'relatorios/fluxo_caixa.html',
        lancamentos_realizado=build_fluxo_rows(lancamentos_realizado, saldo_inicial_por_conta, use_valor_real=False),
        lancamentos_previsto=build_fluxo_rows(lancamentos_previsto, saldo_inicial_por_conta, use_valor_real=True),
        consolidado_realizado=consolidado_realizado,
        consolidado_previsto=consolidado_previsto,
        resumo_diario_realizado=resumo_diario_realizado,
        resumo_diario_previsto=resumo_diario_previsto,
        contas_banco=contas_banco,
        contas_fluxo=contas_fluxo,
        data_inicio=data_inicio,
        data_fim=data_fim,
        conta_banco_id=conta_banco_id,
        conta_fluxo_id=conta_fluxo_id
    )


@relatorios_bp.route('/fluxo-caixa/export-legacy')
@login_required
def export_fluxo_caixa():
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    conta_banco_id = request.args.get('conta_banco_id', '', type=int)
    conta_fluxo_id = request.args.get('conta_fluxo_id', '', type=int)
    if Workbook is None:
        flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
        return redirect(url_for('relatorios.fluxo_caixa'))

    def get_saldo_inicial_por_conta():
        if conta_banco_id:
            contas = ContaBanco.query.filter(
                ContaBanco.empresa_id == current_user.empresa_id,
                ContaBanco.id == conta_banco_id
            ).all()
        else:
            contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
        return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

    def get_saldo_inicial_total():
        return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

    def build_daily_rows(lancamentos, saldo_inicial, use_valor_real, date_attr):
        totals = defaultdict(lambda: {'pagar': Decimal('0.00'), 'receber': Decimal('0.00')})
        for lancamento in lancamentos:
            data_ref = getattr(lancamento, date_attr)
            if not data_ref:
                continue
            valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
            valor = Decimal(str(valor_base or 0))
            if lancamento.fluxo_conta and lancamento.fluxo_conta.is_pagamento():
                totals[data_ref]['pagar'] += valor
            else:
                totals[data_ref]['receber'] += valor

        saldo_atual = saldo_inicial
        rows = []
        for data_ref in sorted(totals.keys()):
            pagar = totals[data_ref]['pagar']
            receber = totals[data_ref]['receber']
            saldo_anterior = saldo_atual
            saldo_atual = saldo_anterior + receber - pagar
            rows.append((data_ref, saldo_anterior, pagar, receber, saldo_atual))
        return rows

    query_realizado = Lancamento.query.filter(
        Lancamento.empresa_id == current_user.empresa_id,
        Lancamento.status == 'pago'
    )
    query_previsto = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)

    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query_realizado = query_realizado.filter(Lancamento.data_pagamento >= data_inicio_dt)
        query_previsto = query_previsto.filter(Lancamento.data_vencimento >= data_inicio_dt)

    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query_realizado = query_realizado.filter(Lancamento.data_pagamento <= data_fim_dt)
        query_previsto = query_previsto.filter(Lancamento.data_vencimento <= data_fim_dt)

    if conta_banco_id:
        query_realizado = query_realizado.filter(Lancamento.conta_banco_id == conta_banco_id)
        query_previsto = query_previsto.filter(Lancamento.conta_banco_id == conta_banco_id)

    if conta_fluxo_id:
        query_realizado = query_realizado.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
        query_previsto = query_previsto.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)

    lancamentos_realizado = query_realizado.order_by(func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc()).all()
    lancamentos_previsto = query_previsto.order_by(func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc()).all()

    saldo_inicial_total = get_saldo_inicial_total()
    resumo_realizado = build_daily_rows(lancamentos_realizado, saldo_inicial_total, False, 'data_pagamento')
    resumo_previsto = build_daily_rows(lancamentos_previsto, saldo_inicial_total, True, 'data_vencimento')

    wb = Workbook()
    ws_previsto = wb.active
    ws_previsto.title = 'Previsto'
    ws_realizado = wb.create_sheet('Realizado')

    headers = ['Data', 'Saldo Anterior', 'Pagamentos', 'Recebimentos', 'Saldo do Dia']
    ws_previsto.append(headers)
    ws_realizado.append(headers)

    for data_ref, saldo_anterior, pagar, receber, saldo_atual in resumo_previsto:
        ws_previsto.append([data_ref.strftime('%d/%m/%Y'), float(saldo_anterior), float(pagar), float(receber), float(saldo_atual)])

    for data_ref, saldo_anterior, pagar, receber, saldo_atual in resumo_realizado:
        ws_realizado.append([data_ref.strftime('%d/%m/%Y'), float(saldo_anterior), float(pagar), float(receber), float(saldo_atual)])

    for sheet in (ws_previsto, ws_realizado):
        for col in ['B', 'C', 'D', 'E']:
            for cell in sheet[col]:
                cell.number_format = '#,##0.00'
        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 16
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='fluxo_caixa_diario.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )