from __future__ import annotations

import io
import logging
import traceback
from calendar import monthrange
from datetime import date, datetime

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required

from src.models import ContaBanco, Entidade
from src.services.fluxo_relatorio import gerar_relatorio_fluxo
from src.tenant import scoped_query, tenant_id

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None
    Font = None
    PatternFill = None


relatorio_fluxo_bp = Blueprint('relatorio_fluxo', __name__, url_prefix='/relatorios/fluxo-caixa')


def _periodo_padrao():
    hoje = date.today()
    inicio = date(hoje.year, hoje.month, 1)
    fim = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])
    return inicio, fim


def _parse_date(value: str, fallback: date) -> date:
    if not value:
        return fallback
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return fallback


@relatorio_fluxo_bp.route('/')
@login_required
def index():
    inicio_padrao, fim_padrao = _periodo_padrao()
    data_inicio = _parse_date(request.args.get('data_inicio', ''), inicio_padrao)
    data_fim = _parse_date(request.args.get('data_fim', ''), fim_padrao)

    tipo_raw = (request.args.get('tipo') or '').strip().upper()
    tipo = tipo_raw if tipo_raw in {'R', 'P'} else None
    conta_banco_id = request.args.get('conta_banco_id', type=int)
    entidade_id = request.args.get('entidade_id', type=int)
    agrupar_por = (request.args.get('agrupar_por') or 'dia').strip().lower()
    if agrupar_por not in {'dia', 'semana', 'mes'}:
        agrupar_por = 'dia'

    formato = (request.args.get('formato') or 'html').strip().lower()

    try:
        resultado = gerar_relatorio_fluxo(
            empresa_id=tenant_id() or current_user.empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            tipo=tipo,
            conta_banco_id=conta_banco_id,
            entidade_id=entidade_id,
            agrupar_por=agrupar_por,
        )
    except Exception as exc:
        logging.error('Erro ao gerar relatorio de fluxo: %s\n%s', exc, traceback.format_exc())
        if formato == 'json':
            return jsonify({'erro': 'Falha ao gerar relatório.'}), 500
        flash('Erro ao gerar relatório de fluxo de caixa.', 'danger')
        return redirect(url_for('dashboard.index'))

    if formato == 'json':
        return jsonify(resultado)

    contas_banco = scoped_query(ContaBanco).filter_by(ativo=True).order_by(ContaBanco.nome.asc()).all()
    entidades = scoped_query(Entidade).filter_by(ativo=True).order_by(Entidade.nome.asc()).all()

    return render_template(
        'relatorios/fluxo_caixa.html',
        resultado=resultado,
        contas_banco=contas_banco,
        entidades=entidades,
        data_inicio=data_inicio.strftime('%Y-%m-%d'),
        data_fim=data_fim.strftime('%Y-%m-%d'),
        tipo=tipo or '',
        conta_banco_id=conta_banco_id,
        entidade_id=entidade_id,
        agrupar_por=agrupar_por,
    )


@relatorio_fluxo_bp.route('/exportar')
@login_required
def exportar():
    if Workbook is None:
        flash('Exportação para Excel indisponível: openpyxl não instalado.', 'warning')
        return redirect(url_for('relatorio_fluxo.index'))

    inicio_padrao, fim_padrao = _periodo_padrao()
    data_inicio = _parse_date(request.args.get('data_inicio', ''), inicio_padrao)
    data_fim = _parse_date(request.args.get('data_fim', ''), fim_padrao)

    tipo_raw = (request.args.get('tipo') or '').strip().upper()
    tipo = tipo_raw if tipo_raw in {'R', 'P'} else None
    conta_banco_id = request.args.get('conta_banco_id', type=int)
    entidade_id = request.args.get('entidade_id', type=int)
    agrupar_por = (request.args.get('agrupar_por') or 'dia').strip().lower()
    if agrupar_por not in {'dia', 'semana', 'mes'}:
        agrupar_por = 'dia'

    resultado = gerar_relatorio_fluxo(
        empresa_id=tenant_id() or current_user.empresa_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tipo=tipo,
        conta_banco_id=conta_banco_id,
        entidade_id=entidade_id,
        agrupar_por=agrupar_por,
    )

    wb = Workbook()
    ws_diario = wb.active
    ws_diario.title = 'Diário'
    ws_resumo = wb.create_sheet('Resumo')
    ws_grafico = wb.create_sheet('Gráfico')

    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    neg_font = Font(color='C00000')

    ws_diario.append(['Data', 'Código', 'Descrição', 'Tipo', 'Previsto', 'Realizado', 'Disponível', 'Desvio', 'Desvio %'])
    for cell in ws_diario[1]:
        cell.fill = header_fill
        cell.font = header_font

    for linha in resultado['linhas']:
        ws_diario.append([
            linha['data'],
            linha['codigo'],
            linha['descricao'],
            linha['tipo'],
            linha['previsto'],
            linha['realizado'],
            linha['disponivel'],
            linha['desvio'],
            linha['desvio_pct'] / 100.0,
        ])

    for row in ws_diario.iter_rows(min_row=2, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = neg_font

    for cell in ws_diario['I'][1:]:
        cell.number_format = '0.00%'
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = neg_font

    ws_resumo.append(['Indicador', 'Valor'])
    for cell in ws_resumo[1]:
        cell.fill = header_fill
        cell.font = header_font

    for key, value in resultado['totais'].items():
        ws_resumo.append([key, value])
    for row in ws_resumo.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = neg_font

    ws_grafico.append(['Data', 'Previsto', 'Realizado', 'Disponível'])
    for cell in ws_grafico[1]:
        cell.fill = header_fill
        cell.font = header_font

    for data_ref, valores in sorted(resultado['por_data'].items()):
        ws_grafico.append([data_ref, valores['previsto'], valores['realizado'], valores['disponivel']])

    for row in ws_grafico.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = neg_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='relatorio_fluxo_caixa.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
