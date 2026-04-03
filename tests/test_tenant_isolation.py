import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from src.app import create_app
from src.models import db, Empresa, User, Entidade, FluxoContaModel, ContaBanco, Lancamento, ImportacaoNFSe, Comissao


class TenantIsolationTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app('testing')
        cls.app.config['WTF_CSRF_ENABLED'] = False
        cls.ctx = cls.app.app_context()
        cls.ctx.push()

        db.drop_all()
        db.create_all()

        empresa_a = Empresa(nome='Empresa A', cnpj='11111111000111')
        empresa_b = Empresa(nome='Empresa B', cnpj='22222222000122')
        db.session.add_all([empresa_a, empresa_b])
        db.session.flush()

        user_a = User(
            empresa_id=empresa_a.id,
            username='financeiro',
            email='user_a@test.local',
            full_name='User A',
            is_active=True,
            is_admin=True,
        )
        user_a.set_password('123456')

        user_b = User(
            empresa_id=empresa_b.id,
            username='financeiro',
            email='user_b@test.local',
            full_name='User B',
            is_active=True,
            is_admin=True,
        )
        user_b.set_password('654321')
        db.session.add_all([user_a, user_b])
        db.session.flush()

        fluxo_a = FluxoContaModel(empresa_id=empresa_a.id, codigo='1.1', descricao='Receita A', tipo='R', ativo=True)
        fluxo_b = FluxoContaModel(empresa_id=empresa_b.id, codigo='1.1', descricao='Receita B', tipo='R', ativo=True)
        fluxo_pagamento_a = FluxoContaModel(empresa_id=empresa_a.id, codigo='2.1', descricao='Despesa A', tipo='P', ativo=True)
        db.session.add_all([fluxo_a, fluxo_b, fluxo_pagamento_a])
        db.session.flush()

        conta_a = ContaBanco(
            empresa_id=empresa_a.id,
            nome='Banco A',
            banco='A',
            agencia='0001',
            numero_conta='12345',
            ativo=True,
            saldo_inicial=1000,
            fluxo_conta_id=fluxo_a.id,
        )
        conta_b = ContaBanco(
            empresa_id=empresa_b.id,
            nome='Banco B',
            banco='B',
            agencia='0002',
            numero_conta='67890',
            ativo=True,
            saldo_inicial=2000,
            fluxo_conta_id=fluxo_b.id,
        )
        db.session.add_all([conta_a, conta_b])
        db.session.flush()

        entidade_a = Entidade(
            empresa_id=empresa_a.id,
            tipo='C',
            cnpj_cpf='00000000000001',
            nome='Cliente A',
            ativo=True,
        )
        entidade_b = Entidade(
            empresa_id=empresa_b.id,
            tipo='C',
            cnpj_cpf='00000000000002',
            nome='Cliente B',
            ativo=True,
        )
        db.session.add_all([entidade_a, entidade_b])
        db.session.flush()

        lancamento_a = Lancamento(
            empresa_id=empresa_a.id,
            data_evento=date(2026, 3, 1),
            data_vencimento=date(2026, 3, 2),
            data_pagamento=date(2026, 3, 2),
            status='pago',
            fluxo_conta_id=fluxo_a.id,
            conta_banco_id=conta_a.id,
            entidade_id=entidade_a.id,
            valor_real=100,
            valor_pago=100,
            numero_documento='DOC-A',
        )
        lancamento_aberto_a = Lancamento(
            empresa_id=empresa_a.id,
            data_evento=date(2026, 3, 10),
            data_vencimento=date(2026, 3, 11),
            data_pagamento=None,
            status='aberto',
            fluxo_conta_id=fluxo_a.id,
            conta_banco_id=conta_a.id,
            entidade_id=entidade_a.id,
            valor_real=300,
            valor_pago=0,
            numero_documento='DOC-ABERTO-A',
        )
        lancamento_saida_a = Lancamento(
            empresa_id=empresa_a.id,
            data_evento=date(2026, 3, 12),
            data_vencimento=date(2026, 3, 13),
            data_pagamento=date(2026, 3, 13),
            status='pago',
            fluxo_conta_id=fluxo_pagamento_a.id,
            conta_banco_id=conta_a.id,
            entidade_id=entidade_a.id,
            valor_real=75,
            valor_pago=75,
            numero_documento='DOC-SAIDA-A',
        )
        lancamento_b = Lancamento(
            empresa_id=empresa_b.id,
            data_evento=date(2026, 3, 1),
            data_vencimento=date(2026, 3, 2),
            data_pagamento=date(2026, 3, 2),
            status='pago',
            fluxo_conta_id=fluxo_b.id,
            conta_banco_id=conta_b.id,
            entidade_id=entidade_b.id,
            valor_real=250,
            valor_pago=250,
            numero_documento='DOC-B',
        )
        db.session.add_all([lancamento_a, lancamento_aberto_a, lancamento_saida_a, lancamento_b])
        importacao_a = ImportacaoNFSe(
            empresa_id=empresa_a.id,
            chave_nota='NFSE-A',
            numero_nota='NFE-A',
            data_emissao=date(2026, 3, 3),
            cnpj_tomador=entidade_a.cnpj_cpf,
            entidade_id=entidade_a.id,
            lancamento_id=lancamento_a.id,
            valor_bruto=100,
            valor_impostos=10,
            descricao_servico='Servico A',
            aliquota_iss=6,
        )
        importacao_b = ImportacaoNFSe(
            empresa_id=empresa_b.id,
            chave_nota='NFSE-B',
            numero_nota='NFE-B',
            data_emissao=date(2026, 3, 4),
            cnpj_tomador=entidade_b.cnpj_cpf,
            entidade_id=entidade_b.id,
            lancamento_id=lancamento_b.id,
            valor_bruto=200,
            valor_impostos=20,
            descricao_servico='Servico B',
            aliquota_iss=6,
        )
        db.session.add_all([importacao_a, importacao_b])
        db.session.commit()

        cls.empresa_a_id = empresa_a.id
        cls.empresa_b_id = empresa_b.id
        cls.user_a_id = user_a.id
        cls.conta_b_id = conta_b.id
        cls.fluxo_pagamento_a_id = fluxo_pagamento_a.id
        cls.entidade_b_id = entidade_b.id
        cls.entidade_a_id = entidade_a.id
        cls.lancamento_a_id = lancamento_a.id
        cls.lancamento_aberto_a_id = lancamento_aberto_a.id
        cls.lancamento_saida_a_id = lancamento_saida_a.id
        cls.lancamento_b_id = lancamento_b.id

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        db.engine.dispose()
        cls.ctx.pop()

    def setUp(self):
        self.client = self.app.test_client()

    def _login_as_user_a(self):
        response = self.client.post(
            '/auth/login',
            data={
                'empresa_cnpj': '11111111000111',
                'username': 'financeiro',
                'password': '123456',
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

    def test_same_username_allowed_in_different_companies(self):
        users = User.query.filter_by(username='financeiro').all()

        self.assertEqual(len(users), 2)

    def test_login_uses_company_scope(self):
        response = self.client.post(
            '/auth/login',
            data={
                'empresa_cnpj': '22222222000122',
                'username': 'financeiro',
                'password': '654321',
            },
            follow_redirects=True,
        )
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('Dashboard', body)
        self.assertNotIn('Empresa, usuário ou senha inválidos', body)

    def test_protected_routes_redirect_anonymous_users(self):
        app = create_app('testing')
        with app.app_context():
            with app.test_client() as client:
                response_comissoes = client.get('/comissoes/', follow_redirects=False)
                response_notas_nfse = client.get('/relatorios/notas_nfse', follow_redirects=False)
                response_importacoes = client.get('/importacoes/nfse', follow_redirects=False)
                response_export_lanc = client.get('/relatorios/lancamentos/export?formato=xlsx', follow_redirects=False)
                response_export_fluxo_csv = client.get('/relatorios/fluxo-caixa-csv/export', follow_redirects=False)
                response_export_comissoes = client.get('/comissoes/exportar-csv', follow_redirects=False)

            db.session.remove()
            db.engine.dispose()

        self.assertEqual(response_comissoes.status_code, 302)
        self.assertEqual(response_notas_nfse.status_code, 302)
        self.assertEqual(response_importacoes.status_code, 302)
        self.assertEqual(response_export_lanc.status_code, 302)
        self.assertEqual(response_export_fluxo_csv.status_code, 302)
        self.assertEqual(response_export_comissoes.status_code, 302)

    def test_entidades_list_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/entidades/', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('Cliente A', body)
        self.assertNotIn('Cliente B', body)

    def test_entidade_detail_blocks_cross_company_access(self):
        self._login_as_user_a()

        response = self.client.get(f'/entidades/{self.entidade_b_id}/ver')

        self.assertEqual(response.status_code, 404)

    def test_contas_banco_list_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/contas-banco/', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('Banco A', body)
        self.assertNotIn('Banco B', body)

    def test_conta_banco_detail_blocks_cross_company_access(self):
        self._login_as_user_a()

        response = self.client.get(f'/contas-banco/{self.conta_b_id}/ver')

        self.assertEqual(response.status_code, 404)

    def test_lancamento_edit_blocks_cross_company_access(self):
        self._login_as_user_a()

        response = self.client.get(f'/lancamentos/{self.lancamento_b_id}/editar')

        self.assertEqual(response.status_code, 404)

    def test_fluxo_csv_does_not_include_other_company_data(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa-csv')
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-A', body)
        self.assertNotIn('DOC-B', body)

    def test_fluxo_csv_filters_by_date(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa-csv?data_fim=2026-03-11', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-A', body)
        self.assertIn('DOC-ABERTO-A', body)
        self.assertNotIn('DOC-SAIDA-A', body)

    def test_fluxo_caixa_export_xlsx_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa/export', follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook['Previsto']
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('100', exported_text)
        self.assertNotIn('250', exported_text)

    def test_fluxo_csv_export_xlsx_filters_by_date(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa-csv/export?data_fim=2026-03-11', follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook.active
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('DOC-A', exported_text)
        self.assertIn('DOC-ABERTO-A', exported_text)
        self.assertNotIn('DOC-SAIDA-A', exported_text)

    def test_fluxo_caixa_csv_export_xlsx_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa-csv/export', follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook.active
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('DOC-A', exported_text)
        self.assertNotIn('DOC-B', exported_text)

    def test_fluxo_caixa_report_filters_by_date(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/fluxo-caixa?data_fim=2026-03-12', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-A', body)
        self.assertIn('DOC-ABERTO-A', body)
        self.assertNotIn('DOC-SAIDA-A', body)

    def test_fluxo_caixa_report_filters_by_conta_banco(self):
        self._login_as_user_a()

        response = self.client.get(f'/relatorios/fluxo-caixa?conta_banco_id={self.conta_b_id}', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('DOC-A', body)
        self.assertNotIn('DOC-ABERTO-A', body)
        self.assertNotIn('DOC-B', body)

    def test_lancamentos_export_xlsx_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/lancamentos/export?formato=xlsx', follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook.active
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('DOC-A', exported_text)
        self.assertNotIn('DOC-B', exported_text)

    def test_lancamentos_report_filters_by_status(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/lancamentos?status=aberto', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-ABERTO-A', body)
        self.assertNotIn('DOC-B', body)

    def test_lancamentos_report_filters_by_tipo_movimento(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/lancamentos?tipo_movimento=P', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-SAIDA-A', body)
        self.assertNotIn('DOC-A', body)
        self.assertNotIn('DOC-ABERTO-A', body)

    def test_lancamentos_report_filters_by_conta_fluxo(self):
        self._login_as_user_a()

        response = self.client.get(f'/relatorios/lancamentos?conta_fluxo_id={self.fluxo_pagamento_a_id}', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-SAIDA-A', body)
        self.assertNotIn('DOC-B', body)

    def test_lancamentos_report_filters_by_vencimento(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/lancamentos?data_venc_ate=2026-03-02', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('DOC-A', body)
        self.assertNotIn('DOC-ABERTO-A', body)

    def test_lancamentos_export_pdf_returns_file(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/lancamentos/export?formato=pdf', follow_redirects=False)

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/pdf', response.content_type)
        self.assertIn('listagem_lancamentos.pdf', response.headers.get('Content-Disposition', ''))

    def test_notas_nfse_report_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/notas_nfse', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('NFE-A', body)
        self.assertNotIn('NFE-B', body)

    def test_notas_nfse_report_filters_by_emissao_date(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/notas_nfse?data_emissao_ate=2026-03-03', follow_redirects=True)
        body = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('NFE-A', body)
        self.assertNotIn('NFE-B', body)

    def test_notas_nfse_export_xlsx_isolated_by_company(self):
        self._login_as_user_a()

        response = self.client.get('/relatorios/notas_nfse/export?formato=xlsx', follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook.active
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('NFE-A', exported_text)
        self.assertNotIn('NFE-B', exported_text)

    def test_notas_nfse_export_xlsx_filters_by_emissao_date(self):
        self._login_as_user_a()

        response = self.client.get(
            '/relatorios/notas_nfse/export?formato=xlsx&data_emissao_ate=2026-03-03',
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        worksheet = workbook.active
        exported_text = '\n'.join(
            str(cell.value or '')
            for row in worksheet.iter_rows()
            for cell in row
        )

        self.assertIn('NFE-A', exported_text)
        self.assertNotIn('NFE-B', exported_text)

    def test_comissoes_export_csv_isolated_by_company(self):
        vendedor_a = Entidade(
            empresa_id=self.empresa_a_id,
            tipo='V',
            cnpj_cpf='00000000000011',
            nome='Vendedor A',
            ativo=True,
        )
        vendedor_b = Entidade(
            empresa_id=self.empresa_b_id,
            tipo='V',
            cnpj_cpf='00000000000022',
            nome='Vendedor B',
            ativo=True,
        )
        db.session.add_all([vendedor_a, vendedor_b])
        db.session.flush()

        comissao_a = Comissao(
            empresa_id=self.empresa_a_id,
            id_apuracao=1,
            lancamento_id=self.lancamento_a_id,
            entidade_cliente_id=self.entidade_a_id,
            entidade_vendedor_id=vendedor_a.id,
            dt_lancamento=date(2026, 3, 1),
            dt_vencimento=date(2026, 3, 2),
            dt_pagamento_recebimento=date(2026, 3, 2),
            vl_nota=100,
            vl_imposto=0,
            vl_outros_custos=0,
            vl_repasse=0,
            vl_liquido=100,
            aliquota_aplicada=10,
            vl_comissao=10,
            situacao='ativo',
        )
        comissao_b = Comissao(
            empresa_id=self.empresa_b_id,
            id_apuracao=1,
            lancamento_id=self.lancamento_b_id,
            entidade_cliente_id=self.entidade_b_id,
            entidade_vendedor_id=vendedor_b.id,
            dt_lancamento=date(2026, 3, 1),
            dt_vencimento=date(2026, 3, 2),
            dt_pagamento_recebimento=date(2026, 3, 2),
            vl_nota=250,
            vl_imposto=0,
            vl_outros_custos=0,
            vl_repasse=0,
            vl_liquido=250,
            aliquota_aplicada=10,
            vl_comissao=25,
            situacao='ativo',
        )
        db.session.add_all([comissao_a, comissao_b])
        db.session.commit()

        self._login_as_user_a()
        response = self.client.get('/comissoes/exportar-csv', follow_redirects=True)
        body = response.data.decode('utf-8-sig')

        self.assertEqual(response.status_code, 200)
        self.assertIn('Vendedor A', body)
        self.assertNotIn('Vendedor B', body)


if __name__ == '__main__':
    unittest.main()
