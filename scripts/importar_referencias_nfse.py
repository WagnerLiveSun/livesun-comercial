from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.app import create_app
from src.models import (
    db,
    NfseIndOpReferencia,
    NfseMunicipioReferencia,
    NfseNbsReferencia,
    NfseServicoNacionalReferencia,
)


DEFAULT_ANEXOS = ROOT / 'Anexos'


def _text(value) -> str:
    return '' if value is None else str(value).strip()


def _digits(value) -> str:
    return ''.join(ch for ch in _text(value) if ch.isdigit())


def _reset_table(model) -> None:
    model.query.delete(synchronize_session=False)


def importar_municipios(anexos_dir: Path) -> int:
    workbook = load_workbook(anexos_dir / 'anexo_a-municipio_ibge-paises_iso2-v1-00-snnfse-20251210.xlsx', read_only=True, data_only=True)
    sheet = workbook['TAB.MUN_IBGE']
    _reset_table(NfseMunicipioReferencia)
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome_uf, uf_sigla, nome_municipio, codigo_ibge = row[:4]
        codigo = _digits(codigo_ibge)
        if not codigo:
            continue
        db.session.add(
            NfseMunicipioReferencia(
                codigo_ibge=codigo.zfill(7)[-7:],
                nome_uf=_text(nome_uf),
                uf_sigla=_text(uf_sigla)[:2],
                nome_municipio=_text(nome_municipio),
                pais_iso2='BR',
                pais_nome='Brasil',
                ativo=True,
            )
        )
        total += 1
    return total


def importar_servicos(anexos_dir: Path) -> int:
    workbook = load_workbook(anexos_dir / 'anexo_b-nbs2-lista_servico_nacional-snnfse-v1-01-20260122.xlsx', read_only=True, data_only=True)
    sheet = workbook['LISTA.SERV.NAC.']
    _reset_table(NfseServicoNacionalReferencia)
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        codigo, item, subitem, desdobro, descricao = row[:5]
        descricao_texto = _text(descricao)
        if not descricao_texto:
            continue
        codigo_normalizado = _digits(codigo)
        db.session.add(
            NfseServicoNacionalReferencia(
                codigo_tributacao_nacional=codigo_normalizado or None,
                item=_text(item) or None,
                subitem=_text(subitem) or None,
                desdobro_nacional=_text(desdobro) or None,
                descricao=descricao_texto,
                ativo=True,
            )
        )
        total += 1
    return total


def importar_nbs(anexos_dir: Path) -> int:
    workbook = load_workbook(anexos_dir / 'anexo_b-nbs2-lista_servico_nacional-snnfse-v1-01-20260122.xlsx', read_only=True, data_only=True)
    sheet = workbook['LISTA.NBS_v2.0']
    _reset_table(NfseNbsReferencia)
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        codigo, descricao = row[:2]
        descricao_texto = _text(descricao)
        if not descricao_texto:
            continue
        codigo_normalizado = _digits(codigo)
        if not codigo_normalizado:
            continue
        db.session.add(
            NfseNbsReferencia(
                codigo_nbs=codigo_normalizado,
                descricao=descricao_texto,
                origem_catalogo='ANEXO_B',
                ativo=True,
            )
        )
        total += 1
    return total


def importar_indop(anexos_dir: Path) -> int:
    workbook = load_workbook(anexos_dir / 'anexo_c-indop_ibscbs-snnfse-v1-01-20260122.xlsx', read_only=True, data_only=True)
    sheet = workbook['IndOp']
    _reset_table(NfseIndOpReferencia)
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        art, tipo_operacao, considera_local_operacao, caracteristica_fornecimento, codigo_0201, codigo_0202, codigo_indop, local_fornecimento_identificado, campo_leiaute = row[:9]
        codigo = _digits(codigo_indop)
        if not codigo:
            continue
        db.session.add(
            NfseIndOpReferencia(
                codigo_indop=codigo,
                art=_text(art) or None,
                tipo_operacao=_text(tipo_operacao),
                considera_local_operacao=_text(considera_local_operacao) or None,
                caracteristica_fornecimento=_text(caracteristica_fornecimento) or None,
                local_fornecimento_identificado=_text(local_fornecimento_identificado) or None,
                campo_leiaute=_text(campo_leiaute) or None,
                ativo=True,
            )
        )
        total += 1
    return total


def main() -> int:
    parser = argparse.ArgumentParser(description='Importa catálogos de referência NFS-e a partir dos XLSX de Anexos.')
    parser.add_argument('--anexos-dir', type=Path, default=DEFAULT_ANEXOS, help='Diretório com os arquivos XLSX.')
    parser.add_argument('--skip-indop', action='store_true', help='Não importa a tabela de IndOp.')
    args = parser.parse_args()

    app = create_app()
    with app.app_context():
        db.create_all()
        total_municipios = importar_municipios(args.anexos_dir)
        total_servicos = importar_servicos(args.anexos_dir)
        total_nbs = importar_nbs(args.anexos_dir)
        total_indop = 0
        if not args.skip_indop:
            total_indop = importar_indop(args.anexos_dir)
        db.session.commit()

    print(f'Municipios importados: {total_municipios}')
    print(f'Servicos importados: {total_servicos}')
    print(f'NBS importados: {total_nbs}')
    print(f'IndOp importados: {total_indop}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
